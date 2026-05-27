"""Tests for file upload helpers — _file_to_dataframe, _infer_column_types,
_dataframe_to_sqlite, and the upload-preview/upload-csv endpoints."""

from __future__ import annotations

import io
from typing import TYPE_CHECKING

import openpyxl
import pandas as pd
import pytest
from fastapi.testclient import TestClient

from insightxpert_api.routes.databases import (
    _dataframe_to_sqlite,
    _file_to_dataframe,
    _infer_column_types,
)

if TYPE_CHECKING:
    pass


# ── _file_to_dataframe ──────────────────────────────────────────────────


def test_file_to_dataframe_csv_utf8():
    df, meta = _file_to_dataframe(b"a,b\n1,2\n3,4\n", "test.csv")
    assert len(df) == 2
    assert list(df.columns) == ["a", "b"]
    assert meta["encoding"] == "utf-8"
    assert meta["original_rows"] == 2


def test_file_to_dataframe_csv_latin1():
    csv = "näme,scöre\nJürgen,95\n".encode("latin-1")
    df, meta = _file_to_dataframe(csv, "latin1.csv")
    assert len(df) == 1
    assert meta["encoding"] == "latin-1"


def test_file_to_dataframe_csv_cp1252():
    csv = "name,value\nJos\xe9,100\n".encode("cp1252")
    df, meta = _file_to_dataframe(csv, "cp1252.csv")
    assert len(df) == 1
    # latin-1 parses cp1252 byte sequences without error, so it may be
    # detected as latin-1 before cp1252 is tried.
    assert meta["encoding"] in ("latin-1", "cp1252")


def test_file_to_dataframe_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "product", "price"
    ws["A2"], ws["B2"] = "Widget", 9.99
    ws["A3"], ws["B3"] = "Gadget", 14.50
    buf = io.BytesIO()
    wb.save(buf)

    df, meta = _file_to_dataframe(buf.getvalue(), "test.xlsx")
    assert len(df) == 2
    assert list(df.columns)[:2] == ["product", "price"]
    assert meta["sheet_name"] == "Sheet"
    assert meta["sheet_names"] == ["Sheet"]


def test_file_to_dataframe_multi_sheet():
    wb = openpyxl.Workbook()
    wb.active.title = "Data"
    wb.active["A1"] = "x"
    wb.create_sheet("Notes")
    buf = io.BytesIO()
    wb.save(buf)

    df, meta = _file_to_dataframe(buf.getvalue(), "multi.xlsx")
    assert len(meta["sheet_names"]) == 2
    assert meta["sheet_name"] == "Data"
    assert len(df) == 0  # only header, no data rows


def test_file_to_dataframe_empty_csv():
    with pytest.raises(Exception) as exc:
        _file_to_dataframe(b"", "empty.csv")
    assert "no data rows" in str(exc.value.detail)


def test_file_to_dataframe_empty_csv_headers_only():
    """CSV with headers but no data rows -> 0-row DataFrame (not an error)."""
    df, meta = _file_to_dataframe(b"col1,col2\n", "headers.csv")
    assert len(df) == 0
    assert len(df.columns) == 2


def test_file_to_dataframe_xls_rejected():
    with pytest.raises(Exception) as exc:
        _file_to_dataframe(b"dummy", "legacy.xls")
    assert "Legacy .xls" in str(exc.value.detail)
    assert ".xlsx" in str(exc.value.detail)


# NOTE: _file_to_dataframe does NOT validate file extensions — that happens at
# the route layer (upload_preview / upload_csv). The helper falls through to
# the CSV parser for unknown extensions, which may succeed with garbage data
# or raise malformed_csv. Extension validation is tested via the endpoint
# tests below (marked xfail until fresh_db is fixed).


# ── _infer_column_types ─────────────────────────────────────────────────


def test_infer_column_types_integer():
    df = pd.DataFrame({"x": ["1", "2", "3"]})
    types = _infer_column_types(df)
    assert types["x"] == "INTEGER"


def test_infer_column_types_real():
    df = pd.DataFrame({"x": ["1.5", "2.7", "3.0"]})
    types = _infer_column_types(df)
    assert types["x"] == "REAL"


def test_infer_column_types_text():
    df = pd.DataFrame({"x": ["hello", "world", "foo"]})
    types = _infer_column_types(df)
    assert types["x"] == "TEXT"


def test_infer_column_types_boolean():
    df = pd.DataFrame({"x": ["true", "false", "true"]})
    types = _infer_column_types(df)
    assert types["x"] == "INTEGER"  # booleans are stored as INTEGER


def test_infer_column_types_mixed():
    df = pd.DataFrame({"name": ["Alice", "Bob"], "score": ["95", "87"], "active": ["true", "false"]})
    types = _infer_column_types(df)
    assert types["name"] == "TEXT"
    assert types["score"] == "INTEGER"
    assert types["active"] == "INTEGER"


# ── _dataframe_to_sqlite ────────────────────────────────────────────────


def test_dataframe_to_sqlite():
    import sqlite3

    df = pd.DataFrame({"name": ["Alice", "Bob"], "score": [95, 87]})
    blob = _dataframe_to_sqlite(df, "my_db")

    assert len(blob) > 0

    # Verify the SQLite can be deserialized and queried.
    con = sqlite3.connect(":memory:")
    con.deserialize(blob)
    rows = con.execute("SELECT * FROM my_db").fetchall()
    assert len(rows) == 2
    total = con.execute("SELECT SUM(score) FROM my_db").fetchone()
    assert (total[0] or 0) == 182
    con.close()


def test_dataframe_to_sqlite_sanitizes_columns():
    import sqlite3

    df = pd.DataFrame({"col with spaces": ["a"], "123bad": ["b"], "a.b!c": ["c"]})
    blob = _dataframe_to_sqlite(df, "test")

    con = sqlite3.connect(":memory:")
    con.deserialize(blob)
    cols = [
        row[1]
        for row in con.execute("PRAGMA table_info(test)").fetchall()
    ]
    # Column names should be sanitized: no spaces, no leading digits, no special chars.
    assert "col_with_spaces" in cols
    assert all(not c[0].isdigit() for c in cols)
    con.close()


# ── upload-preview endpoint ─────────────────────────────────────────────
# These tests require the fresh_db fixture which is currently broken for all
# tests (alembic is configured for Postgres, not SQLite). Marked xfail until
# the test infrastructure is fixed.


@pytest.mark.xfail(reason="fresh_db fixture broken — alembic configured for Postgres, not SQLite")
def test_upload_preview_csv(user_client):
    client, _user = user_client
    resp = client.post(
        "/api/v1/databases/upload-preview",
        files={"file": ("test.csv", b"name,age\nAlice,30\nBob,25\n", "text/csv")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["row_count"] == 2
    assert len(data["columns"]) == 2
    assert data["encoding"] == "utf-8"
    assert len(data["preview_rows"]) == 2
    assert data["preview_rows"][0]["name"] == "Alice"
    assert data["preview_rows"][0]["age"] == "30"


@pytest.mark.xfail(reason="fresh_db fixture broken — alembic configured for Postgres, not SQLite")
def test_upload_preview_xlsx(user_client):
    client, _user = user_client
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "item", "qty"
    ws["A2"], ws["B2"] = "Apple", 10
    buf = io.BytesIO()
    wb.save(buf)

    resp = client.post(
        "/api/v1/databases/upload-preview",
        files={"file": ("data.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["row_count"] == 1
    assert data["sheet_name"] == "Sheet"
    assert data["sheet_names"] == ["Sheet"]
    assert len(data["columns"]) == 2
    assert len(data["preview_rows"]) == 1
    assert data["preview_rows"][0]["item"] == "Apple"


@pytest.mark.xfail(reason="fresh_db fixture broken — alembic configured for Postgres, not SQLite")
def test_upload_preview_empty_file(user_client):
    client, _user = user_client
    resp = client.post(
        "/api/v1/databases/upload-preview",
        files={"file": ("empty.csv", b"", "text/csv")},
    )
    assert resp.status_code == 400
    assert "no data rows" in resp.json()["detail"]


@pytest.mark.xfail(reason="fresh_db fixture broken — alembic configured for Postgres, not SQLite")
def test_upload_preview_unsupported_type(user_client):
    client, _user = user_client
    resp = client.post(
        "/api/v1/databases/upload-preview",
        files={"file": ("doc.pdf", b"%PDF-1.4\n%...", "application/pdf")},
    )
    assert resp.status_code == 400
    assert "Unsupported file type" in resp.json()["detail"]


# ── upload-csv endpoint with Excel ──────────────────────────────────────


@pytest.mark.xfail(reason="fresh_db fixture broken — alembic configured for Postgres, not SQLite")
def test_upload_csv_xlsx(user_client):
    client, _user = user_client
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "col1", "col2"
    ws["A2"], ws["B2"] = "hello", "world"
    buf = io.BytesIO()
    wb.save(buf)

    resp = client.post(
        "/api/v1/databases/upload-csv",
        data={"db_id": "test_xlsx_upload"},
        files={"file": ("data.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["db_id"] == "test_xlsx_upload"
    assert data["source"] == "uploaded"
    assert data["profile_required"] is True
